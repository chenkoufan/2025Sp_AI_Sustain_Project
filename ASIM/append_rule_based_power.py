from __future__ import annotations

from copy import copy
from pathlib import Path
import shutil

from openpyxl import load_workbook
from openpyxl.comments import Comment


DATA_DIR = Path(__file__).resolve().parent / "ASIM-simulation"
SHEET_NAME = "Lighting Results"
PIR_HEADER = "zonal_pir_ceiling_only_power_W"
HYBRID_HEADER = "occupancy_triggered_rule_based_hybrid_power_W"

CEILING_W_PER_LEVEL = 13.2
TASK_W_PER_LEVEL = 1.5


RULES = {
    "Clear": {
        "pir_ceiling": [6, 6, 6, 6, 1, 1, 2, 2, 1, 1, 1, 2] + [0] * 12,
        "hybrid_ceiling": [0] * 24,
        "hybrid_task": [4] * 4 + [1] * 4 + [1] * 4 + [0] * 4 + [2] * 4 + [0] * 4,
    },
    "Overcast": {
        "pir_ceiling": [10, 9, 9, 10, 8, 8, 8, 9, 8, 7, 7, 8,
                        5, 4, 4, 5, 5, 4, 4, 5, 0, 0, 0, 0],
        "hybrid_ceiling": [5] * 4 + [4] * 4 + [3] * 4 + [0] * 12,
        "hybrid_task": [4] * 8 + [3] * 12 + [0] * 4,
    },
    "Night": {
        "pir_ceiling": [13 if seat % 4 in (0, 3) else 11 for seat in range(24)],
        "hybrid_ceiling": [8] * 24,
        "hybrid_task": [3] * 24,
    },
}


def occupied_test(row: int, seat: int) -> str:
    # Wrapping both sides with semicolons prevents seat 1 matching seat 10 or 21.
    return f'ISNUMBER(SEARCH(";{seat};",";"&$B{row}&";"))'


def zone_max_formula(row: int, levels: list[int], zone: int) -> str:
    start = zone * 8
    terms = [f"IF({occupied_test(row, seat)},{levels[seat]},0)"
             for seat in range(start, start + 8)]
    return f"MAX({','.join(terms)})"


def pir_formula(row: int, levels: list[int]) -> str:
    zone_terms = [zone_max_formula(row, levels, zone) for zone in range(3)]
    return f"={CEILING_W_PER_LEVEL}*({'+'.join(zone_terms)})"


def hybrid_formula(row: int, ceiling_levels: list[int], task_levels: list[int]) -> str:
    zone_terms = [zone_max_formula(row, ceiling_levels, zone) for zone in range(3)]
    task_terms = [f"IF({occupied_test(row, seat)},{task_levels[seat]},0)"
                  for seat in range(24)]
    return (
        f"={CEILING_W_PER_LEVEL}*({'+'.join(zone_terms)})"
        f"+{TASK_W_PER_LEVEL}*({'+'.join(task_terms)})"
    )


def copy_column_style(ws, source_col: int, target_col: int) -> None:
    for row in range(1, ws.max_row + 1):
        source = ws.cell(row, source_col)
        target = ws.cell(row, target_col)
        if source.has_style:
            target._style = copy(source._style)
        if source.number_format:
            target.number_format = source.number_format
        if source.alignment:
            target.alignment = copy(source.alignment)
        if source.protection:
            target.protection = copy(source.protection)


def update_workbook(condition: str) -> Path:
    path = DATA_DIR / f"dataset_{condition}.xlsx"
    backup = DATA_DIR / f"dataset_{condition}.before_rule_based_columns.xlsx"
    if not backup.exists():
        shutil.copy2(path, backup)

    wb = load_workbook(path)
    ws = wb[SHEET_NAME]
    headers = {ws.cell(1, col).value: col for col in range(1, ws.max_column + 1)}
    if "seat_list" not in headers:
        raise ValueError(f"{path.name}: seat_list column is missing")
    if ws.max_row != 76:
        raise ValueError(f"{path.name}: expected 75 data rows, found {ws.max_row - 1}")

    original_last_col = ws.max_column
    pir_col = headers.get(PIR_HEADER, original_last_col + 1)
    hybrid_col = headers.get(HYBRID_HEADER, max(original_last_col + 1, pir_col + 1))
    if pir_col > original_last_col:
        copy_column_style(ws, original_last_col, pir_col)
    if hybrid_col > original_last_col:
        copy_column_style(ws, original_last_col, hybrid_col)

    ws.cell(1, pir_col, PIR_HEADER)
    ws.cell(1, hybrid_col, HYBRID_HEADER)
    ws.cell(1, pir_col).comment = Comment(
        "User-specified zonal PIR ceiling-only rule. For each occupied Zone "
        "(seats 0-7, 8-15, or 16-23), the ceiling level is the maximum seat-specific "
        "requirement; an empty Zone is off. Power = 13.2 W x summed Zone levels.",
        "Codex",
    )
    ws.cell(1, hybrid_col).comment = Comment(
        "User-specified occupancy-triggered rule-based hybrid rule. Each occupied "
        "Zone uses the maximum seat-specific ceiling requirement, while each occupied "
        "seat uses its own task-light level. Power = 13.2 W x summed Zone levels + "
        "1.5 W x summed occupied-seat task levels.",
        "Codex",
    )

    rule = RULES[condition]
    for row in range(2, ws.max_row + 1):
        ws.cell(row, pir_col, pir_formula(row, rule["pir_ceiling"]))
        ws.cell(row, hybrid_col, hybrid_formula(
            row, rule["hybrid_ceiling"], rule["hybrid_task"]
        ))
        ws.cell(row, pir_col).number_format = "0.0"
        ws.cell(row, hybrid_col).number_format = "0.0"

    ws.column_dimensions[ws.cell(1, pir_col).column_letter].width = 34
    ws.column_dimensions[ws.cell(1, hybrid_col).column_letter].width = 50
    if ws.auto_filter.ref:
        ws.auto_filter.ref = f"A1:{ws.cell(ws.max_row, hybrid_col).coordinate}"
    wb.save(path)
    return path


if __name__ == "__main__":
    for weather in ("Clear", "Overcast", "Night"):
        print(update_workbook(weather))
