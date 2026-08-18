from __future__ import annotations

from pathlib import Path
import json
import math

import matplotlib as mpl
import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
from openpyxl import load_workbook
from PIL import Image


ROOT = Path(__file__).resolve().parent
DATA_DIR = ROOT / "ASIM-simulation"
OUT_DIR = ROOT / "ASIM2026_analysis"
FIG_DIR = ROOT / "ASIM2026_figures"
OUT_DIR.mkdir(exist_ok=True)
FIG_DIR.mkdir(exist_ok=True)
FIG_WIDTH_MM = 174  # ASIM two-column text width used for all result figures.

WEATHERS = ["Clear", "Overcast", "Night"]
STRATEGIES = ["Zonal PIR ceiling-only", "Rule-based hybrid", "Optimized hybrid"]
PIR_COL = "zonal_pir_ceiling_only_power_W"
RULE_COL = "occupancy_triggered_rule_based_hybrid_power_W"
CEILING_W = 13.2
TASK_W = 1.5

RULES = {
    "Clear": {
        "pir_ceiling": [6, 6, 6, 6, 1, 1, 2, 2, 1, 1, 1, 2] + [0] * 12,
        "hybrid_ceiling": [0] * 24,
        "hybrid_task": [4] * 4 + [1] * 4 + [1] * 4 + [0] * 4 + [2] * 4 + [0] * 4,
    },
    "Overcast": {
        "pir_ceiling": [10, 9, 9, 10, 8, 8, 8, 9, 8, 7, 7, 8,
                        5, 4, 4, 5, 5, 4, 4, 5, 0, 0, 0, 0],
        "hybrid_ceiling": [5] * 4 + [4] * 4 + [3] * 4 + [0] * 12,
        "hybrid_task": [4] * 8 + [3] * 12 + [0] * 4,
    },
    "Night": {
        "pir_ceiling": [13 if seat % 4 in (0, 3) else 11 for seat in range(24)],
        "hybrid_ceiling": [8] * 24,
        "hybrid_task": [3] * 24,
    },
}

COLORS = {
    "Zonal PIR ceiling-only": "#555555",
    "Rule-based hybrid": "#D55E00",
    "Optimized hybrid": "#0072B2",
}
MARKERS = {
    "Zonal PIR ceiling-only": "o",
    "Rule-based hybrid": "s",
    "Optimized hybrid": "^",
}
LINESTYLES = {
    "Zonal PIR ceiling-only": ":",
    "Rule-based hybrid": "--",
    "Optimized hybrid": "-",
}


def parse_seats(value: object) -> list[int]:
    seats = [int(part) for part in str(value).split(";") if part != ""]
    if any(seat < 0 or seat > 23 for seat in seats):
        raise ValueError(f"Seat ID outside 0-23: {value}")
    return seats


def zone_level_sum(seats: list[int], levels: list[int]) -> int:
    return sum(max([levels[s] for s in seats if s // 8 == zone] or [0])
               for zone in range(3))


def load_weather(weather: str) -> pd.DataFrame:
    path = DATA_DIR / f"dataset_{weather}.xlsx"
    wb = load_workbook(path, data_only=True, read_only=True)
    ws = wb["Lighting Results"]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    frame = pd.DataFrame(rows[1:], columns=rows[0])
    required = {"occupancy_count", "seat_list", "power", PIR_COL, RULE_COL}
    missing = required.difference(frame.columns)
    if missing:
        raise ValueError(f"{path.name} missing columns: {sorted(missing)}")
    if len(frame) != 75:
        raise ValueError(f"{path.name}: expected 75 rows, found {len(frame)}")
    if frame[list(required)].isna().any().any():
        raise ValueError(f"{path.name}: missing required values")
    return frame


def build_long_table() -> pd.DataFrame:
    records: list[dict[str, object]] = []
    for weather in WEATHERS:
        frame = load_weather(weather)
        rule = RULES[weather]
        task_cols = [f"task_{seat:02d}" for seat in range(24)]
        ceiling_cols = ["ceiling_1", "ceiling_2", "ceiling_3"]
        for source_row, row in frame.iterrows():
            seats = parse_seats(row["seat_list"])
            occupancy = int(row["occupancy_count"])
            if occupancy != len(seats):
                raise ValueError(f"{weather} row {source_row + 2}: occupancy/seat mismatch")

            pir_ceiling = CEILING_W * zone_level_sum(seats, rule["pir_ceiling"])
            rule_ceiling = CEILING_W * zone_level_sum(seats, rule["hybrid_ceiling"])
            rule_task = TASK_W * sum(rule["hybrid_task"][seat] for seat in seats)
            optimized_ceiling = CEILING_W * sum(float(row[col]) for col in ceiling_cols)
            optimized_task = TASK_W * sum(float(row[col]) for col in task_cols)

            values = {
                "Zonal PIR ceiling-only": (pir_ceiling, 0.0, float(row[PIR_COL])),
                "Rule-based hybrid": (rule_ceiling, rule_task, float(row[RULE_COL])),
                "Optimized hybrid": (optimized_ceiling, optimized_task, float(row["power"])),
            }
            for strategy, (ceiling, task, total) in values.items():
                if not math.isclose(ceiling + task, total, abs_tol=1e-7):
                    raise ValueError(
                        f"{weather} row {source_row + 2} {strategy}: "
                        f"components {ceiling + task} != total {total}"
                    )
                records.append({
                    "weather": weather,
                    "occupancy_count": occupancy,
                    "layout_index": int(source_row % 5) + 1,
                    "seat_list": row["seat_list"],
                    "active_zones": len({seat // 8 for seat in seats}),
                    "strategy": strategy,
                    "ceiling_power_W": ceiling,
                    "task_power_W": task,
                    "total_power_W": total,
                })
    result = pd.DataFrame(records)
    result["weather"] = pd.Categorical(result["weather"], WEATHERS, ordered=True)
    result["strategy"] = pd.Categorical(result["strategy"], STRATEGIES, ordered=True)
    return result


def make_tables(long: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    overall = (
        long.groupby(["weather", "strategy"], observed=True)
        .agg(
            scenario_count=("total_power_W", "size"),
            mean_power_W=("total_power_W", "mean"),
            sd_power_W=("total_power_W", "std"),
            median_power_W=("total_power_W", "median"),
            min_power_W=("total_power_W", "min"),
            max_power_W=("total_power_W", "max"),
            mean_ceiling_power_W=("ceiling_power_W", "mean"),
            mean_task_power_W=("task_power_W", "mean"),
        )
        .reset_index()
    )
    overall["ceiling_share_pct"] = np.where(
        overall["mean_power_W"] > 0,
        overall["mean_ceiling_power_W"] / overall["mean_power_W"] * 100,
        np.nan,
    )

    by_density = (
        long.groupby(["weather", "strategy", "occupancy_count"], observed=True)
        .agg(
            layout_count=("total_power_W", "size"),
            mean_power_W=("total_power_W", "mean"),
            min_power_W=("total_power_W", "min"),
            max_power_W=("total_power_W", "max"),
            sd_power_W=("total_power_W", "std"),
        )
        .reset_index()
    )
    by_density["range_power_W"] = by_density["max_power_W"] - by_density["min_power_W"]

    variability = (
        by_density.groupby(["weather", "strategy"], observed=True)
        .agg(
            mean_within_density_range_W=("range_power_W", "mean"),
            max_within_density_range_W=("range_power_W", "max"),
        )
        .reset_index()
    )
    max_rows = by_density.loc[
        by_density.groupby(["weather", "strategy"], observed=True)["range_power_W"].idxmax(),
        ["weather", "strategy", "occupancy_count"],
    ].rename(columns={"occupancy_count": "occupancy_at_max_range"})
    variability = variability.merge(max_rows, on=["weather", "strategy"], how="left")

    wide = long.pivot_table(
        index=["weather", "occupancy_count", "layout_index", "seat_list"],
        columns="strategy",
        values="total_power_W",
        observed=True,
    ).reset_index()
    savings_rows: list[dict[str, object]] = []
    for weather in WEATHERS:
        subset = wide[wide["weather"] == weather]
        optimized = subset["Optimized hybrid"]
        for baseline in STRATEGIES[:2]:
            base = subset[baseline]
            diff = base - optimized
            savings_rows.append({
                "weather": weather,
                "baseline": baseline,
                "scenario_count": len(subset),
                "baseline_mean_W": base.mean(),
                "optimized_mean_W": optimized.mean(),
                "absolute_reduction_W": diff.mean(),
                "relative_reduction_pct": (base.mean() - optimized.mean()) / base.mean() * 100,
                "optimized_lower_count": int((diff > 1e-8).sum()),
                "equal_count": int((diff.abs() <= 1e-8).sum()),
                "optimized_higher_count": int((diff < -1e-8).sum()),
            })
    savings = pd.DataFrame(savings_rows)
    return overall, by_density, variability, savings


def style() -> None:
    mpl.rcParams.update({
        "font.family": "sans-serif",
        "font.sans-serif": ["Arial", "Helvetica", "Liberation Sans", "DejaVu Sans"],
        "font.size": 7.5,
        "axes.labelsize": 8,
        "axes.titlesize": 8.5,
        "legend.fontsize": 7,
        "xtick.labelsize": 7,
        "ytick.labelsize": 7,
        "axes.linewidth": 0.7,
        "lines.linewidth": 1.4,
        "lines.markersize": 3.5,
        "pdf.fonttype": 42,
        "ps.fonttype": 42,
        "svg.fonttype": "none",
        "savefig.facecolor": "white",
        "figure.facecolor": "white",
    })


def export(fig: plt.Figure, stem: str) -> None:
    png_path = FIG_DIR / f"{stem}.png"
    fig.savefig(png_path, format="png", dpi=600, facecolor="white", transparent=False)
    fig.savefig(FIG_DIR / f"{stem}.tiff", format="tiff", dpi=600,
                facecolor="white", transparent=False)
    fig.savefig(FIG_DIR / f"{stem}.pdf", format="pdf", facecolor="white", transparent=False)
    fig.savefig(FIG_DIR / f"{stem}.svg", format="svg", facecolor="white", transparent=False)
    with Image.open(png_path) as image:
        rgb = image.convert("RGB")
        rgb.save(png_path, format="PNG", dpi=(600, 600), optimize=True)


def plot_occupancy_trends(by_density: pd.DataFrame) -> None:
    style()
    fig, axes = plt.subplots(1, 3, figsize=(174 / 25.4, 62 / 25.4), layout="constrained")
    for panel, (ax, weather) in enumerate(zip(axes, WEATHERS)):
        subset = by_density[by_density["weather"] == weather]
        for strategy in STRATEGIES:
            data = subset[subset["strategy"] == strategy].sort_values("occupancy_count")
            x = data["occupancy_count"].to_numpy(float)
            y = data["mean_power_W"].to_numpy(float)
            lower = data["min_power_W"].to_numpy(float)
            upper = data["max_power_W"].to_numpy(float)
            ax.fill_between(x, lower, upper, color=COLORS[strategy], alpha=0.10, linewidth=0)
            ax.plot(x, y, color=COLORS[strategy], marker=MARKERS[strategy],
                    linestyle=LINESTYLES[strategy], markevery=2, label=strategy)
        ax.set_title(f"({chr(97 + panel)}) {weather}", loc="left", fontweight="bold")
        ax.set_xlim(1, 15)
        ax.set_ylim(bottom=0)
        ax.set_xticks([1, 3, 5, 7, 9, 11, 13, 15])
        ax.set_xlabel("Occupancy count")
        ax.grid(axis="y", color="#D9D9D9", linewidth=0.5)
        ax.spines[["top", "right"]].set_visible(False)
    axes[0].set_ylabel("Lighting power (W)")
    handles, labels = axes[0].get_legend_handles_labels()
    fig.legend(handles, labels, loc="outside upper center", ncol=3, frameon=False)
    export(fig, "results_power_by_occupancy")
    plt.close(fig)


def plot_allocation_and_savings(overall: pd.DataFrame, savings: pd.DataFrame) -> None:
    style()
    fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(174 / 25.4, 70 / 25.4),
                                  gridspec_kw={"width_ratios": [1.35, 1]}, layout="constrained")
    x = np.arange(len(WEATHERS))
    width = 0.23
    offsets = [-width, 0, width]
    for offset, strategy in zip(offsets, STRATEGIES):
        subset = overall[overall["strategy"] == strategy].set_index("weather").loc[WEATHERS]
        ceiling = subset["mean_ceiling_power_W"].to_numpy()
        task = subset["mean_task_power_W"].to_numpy()
        ax1.bar(x + offset, ceiling, width, color="#8C8C8C", edgecolor="black",
                linewidth=0.55)
        ax1.bar(x + offset, task, width, bottom=ceiling, color="#56B4E9",
                edgecolor="black", linewidth=0.55, hatch="///")
    tick_positions = np.concatenate([x + offset for offset in offsets])
    tick_labels = (["PIR"] * len(WEATHERS) + ["Rule"] * len(WEATHERS) +
                   ["Opt."] * len(WEATHERS))
    order = np.argsort(tick_positions)
    ax1.set_xticks(tick_positions[order], np.asarray(tick_labels)[order])
    for group_x, weather in zip(x, WEATHERS):
        ax1.text(group_x, -0.10, weather, ha="center", va="top",
                 transform=ax1.get_xaxis_transform(), fontweight="bold")
    ax1.set_ylabel("Scenario-mean power (W)")
    ax1.set_title("(a) Power allocation", loc="left", fontweight="bold")
    ax1.set_ylim(bottom=0)
    ax1.grid(axis="y", color="#D9D9D9", linewidth=0.5)
    ax1.spines[["top", "right"]].set_visible(False)
    from matplotlib.patches import Patch
    ax1.legend(
        handles=[Patch(facecolor="#8C8C8C", label="Ceiling"),
                 Patch(facecolor="#56B4E9", hatch="///", label="Task")],
        loc="upper left", frameon=False,
    )

    save_pir = savings[savings["baseline"] == STRATEGIES[0]].set_index("weather").loc[WEATHERS]
    save_rule = savings[savings["baseline"] == STRATEGIES[1]].set_index("weather").loc[WEATHERS]
    b1 = ax2.bar(x - width / 2, save_pir["relative_reduction_pct"], width,
                color="#555555", edgecolor="black", linewidth=0.5,
                label="Vs zonal PIR")
    b2 = ax2.bar(x + width / 2, save_rule["relative_reduction_pct"], width,
                color="#D55E00", edgecolor="black", linewidth=0.5,
                hatch="//", label="Vs rule hybrid")
    ax2.axhline(0, color="black", linewidth=0.7)
    ax2.set_xticks(x, WEATHERS)
    ax2.set_ylabel("Optimized reduction (%)")
    ax2.set_title("(b) Relative to baselines", loc="left", fontweight="bold")
    ax2.grid(axis="y", color="#D9D9D9", linewidth=0.5)
    ax2.spines[["top", "right"]].set_visible(False)
    # Keep the legend away from the zero line and the negative Clear bar.
    ax2.legend(frameon=False, loc="upper right")
    for bars in (b1, b2):
        for bar in bars:
            value = bar.get_height()
            va = "bottom" if value >= 0 else "top"
            offset = 1.5 if value >= 0 else -1.5
            ax2.text(bar.get_x() + bar.get_width() / 2, value + offset, f"{value:.1f}",
                     ha="center", va=va, fontsize=6.5)
    export(fig, "results_allocation_and_savings")
    plt.close(fig)


def plot_density_savings(by_density: pd.DataFrame) -> pd.DataFrame:
    wide = by_density.pivot_table(
        index=["weather", "occupancy_count"], columns="strategy",
        values="mean_power_W", observed=True,
    ).reset_index()
    records = []
    for baseline in STRATEGIES[:2]:
        values = np.where(
            wide[baseline] > 0,
            (wide[baseline] - wide["Optimized hybrid"]) / wide[baseline] * 100,
            np.nan,
        )
        for (_, row), value in zip(wide.iterrows(), values):
            records.append({
                "weather": row["weather"],
                "occupancy_count": row["occupancy_count"],
                "baseline": baseline,
                "optimized_reduction_pct": value,
            })
    density_savings = pd.DataFrame(records)

    style()
    weather_colors = {"Clear": "#0072B2", "Overcast": "#D55E00", "Night": "#009E73"}
    weather_markers = {"Clear": "o", "Overcast": "s", "Night": "^"}
    fig, axes = plt.subplots(1, 2, figsize=(174 / 25.4, 60 / 25.4),
                             sharey=True, layout="constrained")
    titles = ["(a) Relative to zonal PIR", "(b) Relative to rule-based hybrid"]
    for ax, baseline, title in zip(axes, STRATEGIES[:2], titles):
        for weather in WEATHERS:
            data = density_savings[
                (density_savings["baseline"] == baseline) &
                (density_savings["weather"] == weather)
            ].sort_values("occupancy_count")
            ax.plot(data["occupancy_count"], data["optimized_reduction_pct"],
                    color=weather_colors[weather], marker=weather_markers[weather],
                    markevery=2, label=weather)
        ax.axhline(0, color="black", linewidth=0.7)
        ax.set_title(title, loc="left", fontweight="bold")
        ax.set_xlim(1, 15)
        ax.set_xticks([1, 3, 5, 7, 9, 11, 13, 15])
        ax.set_xlabel("Occupancy count")
        ax.grid(axis="y", color="#D9D9D9", linewidth=0.5)
        ax.spines[["top", "right"]].set_visible(False)
    axes[0].set_ylabel("Optimized power reduction (%)")
    axes[1].legend(frameon=False, loc="best")
    export(fig, "results_savings_by_occupancy")
    plt.close(fig)
    return density_savings


def make_report(overall: pd.DataFrame, by_density: pd.DataFrame,
                variability: pd.DataFrame, savings: pd.DataFrame) -> str:
    def row(weather: str, strategy: str) -> pd.Series:
        return overall[(overall.weather == weather) & (overall.strategy == strategy)].iloc[0]

    def saving(weather: str, baseline: str) -> pd.Series:
        return savings[(savings.weather == weather) & (savings.baseline == baseline)].iloc[0]

    lines = [
        "# Comprehensive control-strategy results",
        "",
        "## Analysis scope",
        "",
        "The analysis contains 225 weather–layout cases: three sky conditions, 15 occupancy counts, "
        "and five seat distributions per count. The same seat distributions are compared across all "
        "three strategies, so strategy differences are paired by weather and layout. Values are "
        "instantaneous lighting power (W), not time-integrated energy. The five layouts are design "
        "samples; min–max bands quantify seating-pattern sensitivity and are not confidence intervals.",
        "",
        "## Draft Results text",
        "",
    ]
    paragraphs = []
    for weather in WEATHERS:
        pir = row(weather, STRATEGIES[0])
        rule = row(weather, STRATEGIES[1])
        opt = row(weather, STRATEGIES[2])
        sp = saving(weather, STRATEGIES[0])
        sr = saving(weather, STRATEGIES[1])
        relation = "reduction" if sr.relative_reduction_pct >= 0 else "increase"
        paragraphs.append(
            f"Under {weather.lower()} conditions, scenario-mean power was "
            f"{pir.mean_power_W:.1f} W for zonal PIR ceiling-only control, "
            f"{rule.mean_power_W:.1f} W for rule-based hybrid control, and "
            f"{opt.mean_power_W:.1f} W for optimized hybrid control. Relative to zonal PIR, "
            f"the optimized strategy reduced mean power by {sp.relative_reduction_pct:.1f}% "
            f"({sp.absolute_reduction_W:.1f} W). Relative to the rule-based hybrid strategy, "
            f"it produced a {abs(sr.relative_reduction_pct):.1f}% {relation} "
            f"({abs(sr.absolute_reduction_W):.1f} W in magnitude)."
        )
    lines.extend(paragraphs)

    lines.extend(["", "### Occupancy and layout effects", ""])
    endpoint_parts = []
    for weather in WEATHERS:
        vals = []
        for occ in (1, 15):
            value = by_density[
                (by_density.weather == weather) &
                (by_density.strategy == "Optimized hybrid") &
                (by_density.occupancy_count == occ)
            ].iloc[0]
            vals.append(value.mean_power_W)
        endpoint_parts.append(f"{weather}: {vals[0]:.1f}–{vals[1]:.1f} W")
    lines.append(
        "Mean optimized power increased overall from one to 15 occupants as follows: "
        + "; ".join(endpoint_parts) + ". The non-monotonic steps and shaded min–max bands in the "
        "occupancy plots show that seat location, zone activation, and daylight availability affect "
        "power in addition to occupant count."
    )

    lines.extend(["", "### Seating-pattern sensitivity", ""])
    for weather in WEATHERS:
        parts = []
        for strategy in STRATEGIES:
            v = variability[(variability.weather == weather) & (variability.strategy == strategy)].iloc[0]
            occupant_word = "occupant" if int(v.occupancy_at_max_range) == 1 else "occupants"
            parts.append(
                f"{strategy}: mean within-count range {v.mean_within_density_range_W:.1f} W "
                f"(maximum {v.max_within_density_range_W:.1f} W at "
                f"{int(v.occupancy_at_max_range)} {occupant_word})"
            )
        lines.append(f"- **{weather}:** " + "; ".join(parts) + ".")

    lines.extend(["", "### Power allocation", ""])
    for weather in WEATHERS:
        opt = row(weather, "Optimized hybrid")
        rule = row(weather, "Rule-based hybrid")
        lines.append(
            f"- **{weather}:** optimized control allocated {opt.mean_ceiling_power_W:.1f} W to "
            f"ceiling lighting and {opt.mean_task_power_W:.1f} W to task lighting "
            f"({opt.ceiling_share_pct:.1f}% ceiling). The rule-based hybrid allocation was "
            f"{rule.mean_ceiling_power_W:.1f} W ceiling and {rule.mean_task_power_W:.1f} W task."
        )

    clear_rule = saving("Clear", "Rule-based hybrid")
    lines.extend([
        "",
        "## Interpretation requiring verification",
        "",
        f"The optimized solution was lower than the rule-based hybrid in "
        f"{int(clear_rule.optimized_lower_count)}/75 Clear cases, equal in "
        f"{int(clear_rule.equal_count)}/75, and higher in "
        f"{int(clear_rule.optimized_higher_count)}/75. Consequently, the Clear rule-based hybrid "
        "has a lower scenario-average power than the current GA output. Before describing the GA as "
        "globally energy-optimal, both strategies should be rechecked against identical illuminance "
        "and uniformity constraints. If the rule baseline is feasible under those same constraints, "
        "the result points to optimizer convergence, objective weighting, or search-space settings "
        "rather than an advantage of optimization under Clear sky.",
        "",
        "## Figure captions",
        "",
        "**Power by occupancy.** Mean lighting power for zonal PIR ceiling-only, rule-based hybrid, "
        "and optimized hybrid control under (a) Clear, (b) Overcast, and (c) Night conditions. "
        "Lines show the mean across five seat distributions at each occupancy count; shaded bands "
        "show the observed minimum–maximum range (n = 5 layouts per count). Panel y-axis ranges "
        "differ to retain visibility across sky conditions.",
        "",
        "**Allocation and savings.** (a) Scenario-mean ceiling and task-light power for each strategy "
        "and sky condition, averaged equally across 75 occupancy-layout cases per condition. The "
        "x-axis identifies the PIR, rule-based hybrid, and optimized (Opt.) strategies; fill and "
        "hatching identify lighting components. (b) Percentage "
        "reduction in optimized-hybrid mean power relative to the two rule-based baselines. Negative "
        "values denote higher, rather than lower, optimized power.",
        "",
        "**Savings by occupancy.** Density-resolved optimized-hybrid power reduction relative to "
        "(a) zonal PIR ceiling-only and (b) rule-based hybrid control. Each point is calculated from "
        "the mean power of five matched seat distributions at that occupancy count; the horizontal "
        "line marks no difference.",
        "",
        "## Alt text",
        "",
        "The occupancy figure contains three line-chart panels. Power generally rises with occupancy "
        "and with decreasing daylight. Optimized hybrid remains below zonal PIR in all three conditions, "
        "but the Clear rule-based hybrid line often lies below the optimized line. Shaded ranges are "
        "widest at low-to-intermediate occupancy where seat placement changes which zones are active.",
        "",
        "## Reporting notes",
        "",
        "- Report these quantities as **power (W)**. Energy requires an operating duration and should be "
        "reported in Wh or kWh.",
        "- Scenario means weight every occupancy count and layout equally; they are comparative test-set "
        "summaries, not predictions of annual building energy.",
        "- No inferential significance tests were applied because the five layouts are simulation design "
        "samples rather than independent measurements from a target population.",
    ])
    return "\n".join(lines) + "\n"


def write_manifest() -> None:
    manifest = {
        "analysis_date": "2026-08-11",
        "source_files": [f"ASIM-simulation/dataset_{w}.xlsx" for w in WEATHERS],
        "analysis_unit": "one weather-by-occupancy-by-seat-layout simulation case",
        "replicate_structure": "5 seat distributions for each occupancy count; matched across strategies and weather",
        "transformations": [
            "Workbook cached formula values read with openpyxl data_only=True",
            "Power components recomputed from 13.2 W per ceiling-zone level and 1.5 W per task-light level",
            "At each occupancy count, plotted line is arithmetic mean and band is observed min-max across 5 layouts",
            "Scenario means assign equal weight to all 75 cases per weather condition",
        ],
        "missing_data": "none in required columns",
        "exclusions": "none",
        "uncertainty": "observed min-max layout range; not a confidence interval",
        "figure_width_mm": 174,
        "formats": ["PNG 600 dpi", "PDF vector", "SVG vector/editable text"],
        "software": {
            "python": "3.13 (local Anaconda runtime)",
            "pandas": pd.__version__,
            "numpy": np.__version__,
            "matplotlib": mpl.__version__,
        },
    }
    (OUT_DIR / "analysis_manifest.json").write_text(json.dumps(manifest, indent=2), encoding="utf-8")


def main() -> None:
    long = build_long_table()
    case_rows = long[long["strategy"] == STRATEGIES[0]]
    counts = case_rows.groupby(["weather", "occupancy_count"], observed=True).size()
    if not (counts == 5).all():
        raise ValueError("Expected exactly five layouts for every weather/occupancy combination")
    matched = case_rows.groupby(["occupancy_count", "layout_index"], observed=True)["seat_list"].nunique()
    if not (matched == 1).all():
        raise ValueError("Seat distributions are not matched across weather conditions")
    overall, by_density, variability, savings = make_tables(long)
    density_savings = plot_density_savings(by_density)
    plot_occupancy_trends(by_density)
    plot_allocation_and_savings(overall, savings)

    long.to_csv(OUT_DIR / "control_strategy_results_long.csv", index=False)
    overall.to_csv(OUT_DIR / "overall_power_summary.csv", index=False, float_format="%.6f")
    by_density.to_csv(OUT_DIR / "power_by_occupancy_summary.csv", index=False, float_format="%.6f")
    variability.to_csv(OUT_DIR / "layout_sensitivity_summary.csv", index=False, float_format="%.6f")
    savings.to_csv(OUT_DIR / "optimized_savings_summary.csv", index=False, float_format="%.6f")
    density_savings.to_csv(OUT_DIR / "optimized_savings_by_occupancy.csv", index=False, float_format="%.6f")
    (OUT_DIR / "results_analysis.md").write_text(
        make_report(overall, by_density, variability, savings), encoding="utf-8"
    )
    write_manifest()
    print(f"rows_long={len(long)}")
    print(f"required_missing={int(long.isna().any().sum())}")
    print(overall[["weather", "strategy", "mean_power_W"]].to_string(index=False))
    print(savings.to_string(index=False))


if __name__ == "__main__":
    main()
