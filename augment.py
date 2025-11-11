"""
完整版本：
1) 先完整复制 Excel（保留颜色/格式/列宽/行高等）到 augmented_dataset.xlsx
2) 在复制后的文件基础上，读取 dataset 表
3) 对 Augmented 为空的行进行 2^6 组水平镜像（对 seat/task 成对镜像，ceiling 不变）
4) 对所有行进行上下镜像（对 seat/task 成对镜像，ceiling 上下交换）
5) 去重（基于 seat/task/ceiling 列）
4) 将增强后的行追加到该表末尾，并把 Augmented 标记为 YES
依赖：openpyxl, pandas, numpy
"""

from pathlib import Path
from copy import copy
import pandas as pd
import numpy as np
from openpyxl import load_workbook, Workbook

# ============================================================
# ⚙️ 参数设置
# ============================================================
input_file_name = "dataset" + ".xlsx"
SRC_PATH = Path("dataset/" + input_file_name)
DST_PATH = Path("dataset/augmented_" + input_file_name)
SHEET = "dataset"
APPEND_MODE = False  # False=覆盖写入, True=追加写入

# ============================================================
# Step 1: 完整复制 Excel（保留格式）
# ============================================================
def copy_workbook_with_format(src_path: Path, dst_path: Path):
    wb_src = load_workbook(src_path)
    wb_dst = Workbook()
    wb_dst.remove(wb_dst.active)

    for sheet_name in wb_src.sheetnames:
        ws_src = wb_src[sheet_name]
        ws_dst = wb_dst.create_sheet(sheet_name)

        # 值 + 样式
        for row in ws_src.iter_rows():
            for cell in row:
                new_cell = ws_dst.cell(row=cell.row, column=cell.column, value=cell.value)
                if cell.has_style:
                    new_cell.font = copy(cell.font)
                    new_cell.border = copy(cell.border)
                    new_cell.fill = copy(cell.fill)
                    new_cell.number_format = copy(cell.number_format)
                    new_cell.protection = copy(cell.protection)
                    new_cell.alignment = copy(cell.alignment)

        # 列宽与行高
        for col_letter, dim in ws_src.column_dimensions.items():
            ws_dst.column_dimensions[col_letter].width = dim.width
        for row_idx, dim in ws_src.row_dimensions.items():
            ws_dst.row_dimensions[row_idx].height = dim.height

        # 合并单元格
        for merged in ws_src.merged_cells.ranges:
            ws_dst.merge_cells(str(merged))

        # 页面设置
        ws_dst.sheet_format = copy(ws_src.sheet_format)
        ws_dst.page_margins = copy(ws_src.page_margins)
        ws_dst.page_setup = copy(ws_src.page_setup)

    wb_dst.save(dst_path)

# ============================================================
# Step 2: 定义镜像与增强逻辑
# ============================================================
def row_groups():
    """六个水平组（用于左右镜像）"""
    return [
        (0,1,2,3),
        (4,5,6,7),
        (8,9,10,11),
        (12,13,14,15),
        (16,17,18,19),
        (20,21,22,23),
    ]

def seat_cols():
    return [f"seat_{i:02d}" for i in range(24)]

def task_cols():
    return [f"task_{i:02d}" for i in range(24)]

def ceiling_cols():
    return [f"ceiling_{i}" for i in [1,2,3]]

def ensure_augmented_column_df(df: pd.DataFrame) -> pd.DataFrame:
    if "Augmented" not in df.columns:
        df["Augmented"] = ""
    return df

# ---------- 左右镜像 ----------
def mirror_mapping_for_group(group):
    a,b,c,d = group
    return {a:d, b:c, c:b, d:a}

def mirror_one_group(row_dict, group, seat_names, task_names):
    m = mirror_mapping_for_group(group)
    original = {i: row_dict.get(seat_names[i], None) for i in m.keys()}
    for i,j in m.items():
        row_dict[seat_names[i]] = original[j]
    original_t = {i: row_dict.get(task_names[i], None) for i in m.keys()}
    for i,j in m.items():
        row_dict[task_names[i]] = original_t[j]

def apply_mask_to_row(src_row: pd.Series, mask: int, groups) -> dict:
    seat_names = seat_cols()
    task_names = task_cols()
    rd = src_row.to_dict()
    for bit, grp in enumerate(groups):
        if (mask >> bit) & 1:
            mirror_one_group(rd, grp, seat_names, task_names)
    rd["Augmented"] = "YES"
    return rd

def is_empty_augmented(val) -> bool:
    if val is None:
        return True
    if isinstance(val, float) and np.isnan(val):
        return True
    if isinstance(val, str) and val.strip() == "":
        return True
    return False

# ---------- 上下镜像 ----------
def vertical_mirror_row(row_dict):
    """对单行执行上下镜像"""
    pairs = [
        (0,20),(1,21),(2,22),(3,23),
        (4,16),(5,17),(6,18),(7,19),
        (8,12),(9,13),(10,14),(11,15)
    ]
    seats = seat_cols()
    tasks = task_cols()
    for i,j in pairs:
        row_dict[seats[i]], row_dict[seats[j]] = row_dict[seats[j]], row_dict[seats[i]]
    for i,j in pairs:
        row_dict[tasks[i]], row_dict[tasks[j]] = row_dict[tasks[j]], row_dict[tasks[i]]
    # ceiling 1 ↔ 3
    row_dict["ceiling_1"], row_dict["ceiling_3"] = row_dict["ceiling_3"], row_dict["ceiling_1"]
    return row_dict

# ============================================================
# Step 3: 写入工作表
# ============================================================
def clear_worksheet(ws):
    """清空除表头外的内容"""
    max_row = ws.max_row
    for _ in range(max_row - 1):
        ws.delete_rows(2)
    print("已清空原数据内容，只保留表头。")

def append_rows_to_worksheet(ws, headers, rows_dicts):
    if "id" not in headers:
        headers.insert(0, "id")
        ws.insert_cols(1)
        ws.cell(row=1, column=1, value="id")

    if "Augmented" not in headers:
        headers.append("Augmented")
        ws.cell(row=1, column=len(headers), value="Augmented")

    for rd in rows_dicts:
        ws.append([rd.get(col, None) for col in headers])

# ============================================================
# Step 4: 主流程
# ============================================================
def main():
    print("复制文件并保留格式...")
    copy_workbook_with_format(SRC_PATH, DST_PATH)

    print("读取源数据...")
    df = pd.read_excel(SRC_PATH, sheet_name=SHEET)
    df = ensure_augmented_column_df(df)
    mask_empty = df["Augmented"].apply(is_empty_augmented)
    rows_to_augment = df[mask_empty].copy()

    # 左右镜像增强
    print("生成左右镜像增强数据...")
    groups = row_groups()
    variants_per_row = 2 ** len(groups)
    augmented_rows = []
    for _, row in rows_to_augment.iterrows():
        for m in range(variants_per_row):
            augmented_rows.append(apply_mask_to_row(row, m, groups))
    aug_df = pd.DataFrame(augmented_rows)

    # 上下镜像增强
    print("执行上下镜像...")
    combined_df = pd.concat([df, aug_df], ignore_index=True)
    v_mirrored_rows = []
    for _, row in combined_df.iterrows():
        mirrored = vertical_mirror_row(row.to_dict().copy())
        mirrored["Augmented"] = "YES"
        v_mirrored_rows.append(mirrored)
    v_df = pd.DataFrame(v_mirrored_rows)

    full_df = pd.concat([combined_df, v_df], ignore_index=True)

    # 去重
    print("清理重复行...")
    key_cols = task_cols() + seat_cols() + ceiling_cols()
    before = len(full_df)
    full_df = full_df.drop_duplicates(subset=key_cols, keep="first")
    after = len(full_df)
    print(f"已删除重复行: {before - after}, 剩余: {after}")

    # 添加 id
    full_df.insert(0, "id", range(len(full_df)))

    # 写入 Excel
    print("写入增强数据...")
    wb = load_workbook(DST_PATH)
    ws = wb[SHEET]
    headers = [c.value for c in ws[1] if c.value is not None]

    if not APPEND_MODE:
        clear_worksheet(ws)
    append_rows_to_worksheet(ws, headers, full_df.to_dict(orient="records"))
    wb.save(DST_PATH)

    print(f"✅ 完成! 输出文件: {DST_PATH}")
    print(f"模式: {'追加写入' if APPEND_MODE else '覆盖写入'}, 最终行数: {len(full_df)}")

if __name__ == "__main__":
    main()
